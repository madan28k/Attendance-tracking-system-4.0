from django.shortcuts import render
from django.http import HttpResponseRedirect
from django.urls import reverse
from .models import Student
import qrcode
import socket
from StudentView.views import present
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font


def qrgenerator():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    s.connect(("8.8.8.8", 80))
    ip = s.getsockname()[0]

    link = f"http://{ip}:8000/add_manually"

    # Function to generate and display a QR code
    def generate_qr_code(link):
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(link)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img.save("FacultyView/static/FacultyView/qrcode.png")

    generate_qr_code(link)


def faculty_view(request):
    if request.method == "POST":
        student_roll = request.POST["student_id"]
        student = Student.objects.get(s_roll=student_roll)
        if student in present:
            present.remove(student)
        return HttpResponseRedirect("/")

    else:
        qrgenerator()

        # Convert 'present' set to a list of dictionaries
        request.session['present_students'] = [
            {'s_roll': student.s_roll, 's_fname': student.s_fname, 's_lname': student.s_lname}
            for student in present
        ]
        
        return render(
            request,
            "FacultyView/FacultyViewIndex.html",
            {
                "students": present,
            },
        )


def add_manually(request):
    students = Student.objects.all().order_by("s_roll")
    return render(
        request,
        "StudentView/StudentViewIndex.html",
        {
            "students": students,
        },
    )

def generate_excel(request):
    # Create a workbook and sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Attendance Sheet"

    # Add headers
    headers = ["Student ID", "First Name", "Last Name"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Fetch the present students from the session (as a list of dicts)
    present_students = request.session.get('present_students', [])

    # Add student data
    for row_num, student in enumerate(present_students, 2):
        sheet.cell(row=row_num, column=1, value=student['s_roll'])
        sheet.cell(row=row_num, column=2, value=student['s_fname'])
        sheet.cell(row=row_num, column=3, value=student['s_lname'])

    # Prepare the response with the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="AttendanceSheet.xlsx"'
    wb.save(response)
    return response

